from __future__ import annotations

from pathlib import Path

from django.conf import settings
from django.db import transaction
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from apps.common.choices import DepartmentChoices
from apps.common.events import DomainEvent, event_bus
from apps.reports.events import REPORT_GENERATED
from apps.reports.models import MonitorReportSnapshot
from apps.reports.selectors import aggregate_monitor_metrics, build_monitor_rows_for_user


@transaction.atomic
def generate_monitor_report(*, monitor, start_date, end_date, generated_by=None) -> MonitorReportSnapshot:
    metrics = aggregate_monitor_metrics(monitor=monitor, start_date=start_date, end_date=end_date)
    snapshot, _ = MonitorReportSnapshot.objects.update_or_create(
        monitor=monitor,
        start_date=start_date,
        end_date=end_date,
        defaults={
            "generated_by": generated_by,
            "department": monitor.department,
            "normal_minutes": metrics["normal_minutes"],
            "approved_overtime_minutes": metrics["approved_overtime_minutes"],
            "pending_overtime_minutes": metrics["pending_overtime_minutes"],
            "penalty_minutes": metrics["penalty_minutes"],
            "late_count": metrics["late_count"],
            "annotation_delta_minutes": metrics["annotation_delta_minutes"],
            "total_minutes": metrics["total_minutes"],
            "has_memorandum": metrics["has_memorandum"],
        },
    )
    event_bus.publish(
        DomainEvent(
            name=REPORT_GENERATED,
            aggregate_id=str(snapshot.id),
            payload={
                "report_id": str(snapshot.id),
                "department": snapshot.department,
                "monitor_id": str(snapshot.monitor_id),
            },
        )
    )
    return snapshot


DASHBOARD_EXPORT_DIRECTORY = Path(settings.MEDIA_ROOT) / "dashboard_exports"
DEPARTMENT_EXPORT_FILENAMES = {
    DepartmentChoices.INFORMATICS_LABS: "dashboard_monitores.xlsx",
    DepartmentChoices.PHYSICS: "dashboard_monitores_fisica.xlsx",
    DepartmentChoices.ELECTRICAL: "dashboard_monitores_laboratorios.xlsx",
}


def get_dashboard_export_directory() -> Path:
    DASHBOARD_EXPORT_DIRECTORY.mkdir(parents=True, exist_ok=True)
    return DASHBOARD_EXPORT_DIRECTORY


def export_department_dashboard_to_excel(*, user, department: str) -> Path:
    export_directory = get_dashboard_export_directory()
    file_name = DEPARTMENT_EXPORT_FILENAMES.get(department, f"dashboard_{department}.xlsx")
    export_path = export_directory / file_name
    rows = build_monitor_rows_for_user(user, department=department)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Dashboard"
    headers = [
        "Monitor",
        "Normales (h)",
        "Horas extra aprobadas (h)",
        "Horas extra por aprobar (h)",
        "Anotaciones (h)",
        "Total (h)",
        "Faltan para 192 h",
    ]
    worksheet.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row in rows:
        worksheet.append(
            [
                row["monitor"].full_name,
                row["normal_hours"],
                row["approved_overtime_hours"],
                row["pending_overtime_hours"],
                row["annotation_hours"],
                row["total_hours"],
                row["remaining_hours"],
            ]
        )

    widths = {
        "A": 36,
        "B": 16,
        "C": 24,
        "D": 24,
        "E": 18,
        "F": 14,
        "G": 20,
    }
    for column_letter, width in widths.items():
        worksheet.column_dimensions[column_letter].width = width
    worksheet.freeze_panes = "A2"

    workbook.save(export_path)
    workbook.close()
    return export_path
