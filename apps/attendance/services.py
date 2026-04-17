from __future__ import annotations

import logging
from datetime import datetime
from typing import Optional, Tuple

from django.core.exceptions import ValidationError
from django.core.files.base import File
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from apps.attendance.events import ATTENDANCE_IMPORTED, ATTENDANCE_RECONCILIATION_FAILED
from apps.attendance.models import AttendanceImportJob, AttendanceRawRecord
from apps.attendance.validators import coerce_date, coerce_datetime, resolve_headers, validate_excel_extension, coerce_time
from apps.common.choices import DepartmentChoices, ImportJobStatusChoices, ReconciliationStatusChoices, UserRoleChoices
from apps.common.events import DomainEvent, event_bus
from apps.common.permissions import department_allowed
from apps.common.utils import normalize_text
from apps.monitors.models import Monitor

logger = logging.getLogger(__name__)

DEPARTMENT_LABELS = {
    DepartmentChoices.PHYSICS: "Monitores Fisica",
    DepartmentChoices.INFORMATICS_LABS: "Monitores Aulas de Software",
    DepartmentChoices.ELECTRICAL: "Monitores Laboratorios",
}


def _rewind_uploaded_file(uploaded_file) -> None:
    if hasattr(uploaded_file, "seek"):
        uploaded_file.seek(0)


def _department_label(value: str) -> str:
    return DEPARTMENT_LABELS.get(value, value or "sin dependencia")


def _match_monitor_queryset(*, raw_full_name: str, raw_department: str):
    mapped_department = _map_department(normalize_text(raw_department))
    queryset = Monitor.objects.filter(
        normalized_full_name=normalize_text(raw_full_name),
        is_active=True,
    )
    if mapped_department is not None:
        queryset = queryset.filter(department=mapped_department)
    return queryset


def _preview_workbook_departments(uploaded_file) -> tuple[set[str], dict[int, str]]:
    workbook = None
    try:
        _rewind_uploaded_file(uploaded_file)
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            raise ValidationError("El archivo esta vacio.")

        header_map = resolve_headers([str(item) for item in headers])
        detected_departments: set[str] = set()
        unknown_rows: dict[int, str] = {}

        for row_number, row in enumerate(rows, start=2):
            if not any(value is not None and str(value).strip() for value in row):
                continue

            raw_department = str(row[header_map["department"]] or "").strip()
            mapped_department = _map_department(normalize_text(raw_department))
            if mapped_department is None:
                unknown_rows[row_number] = raw_department or "(vacio)"
                continue
            detected_departments.add(mapped_department)

        return detected_departments, unknown_rows
    except ValidationError:
        raise
    except Exception as exc:
        raise ValidationError(f"No fue posible validar el archivo Excel antes de subirlo: {exc}")
    finally:
        if workbook is not None:
            workbook.close()
        _rewind_uploaded_file(uploaded_file)

def _validate_import_scope(*, uploaded_file, uploaded_by=None) -> None:
    if uploaded_by is None or uploaded_by.role != UserRoleChoices.LEADER:
        return
    if not uploaded_by.department:
        raise ValidationError("El lider no tiene una dependencia configurada para subir registros.")

    detected_departments, unknown_rows = _preview_workbook_departments(uploaded_file)
    if unknown_rows:
        sample_rows = ", ".join(
            f"fila {row_number}: {value}" for row_number, value in sorted(unknown_rows.items())[:3]
        )
    
        raise ValidationError(
            "No se pudo validar la dependencia del archivo. "
            f"Revisa la columna Departamento. Valores no reconocidos: {sample_rows}."
        )

    foreign_departments = sorted(
        department for department in detected_departments if department != uploaded_by.department
    )
    if foreign_departments:
        foreign_labels = ", ".join(_department_label(department) for department in foreign_departments)
        raise ValidationError(
            f"Solo puedes subir registros de {_department_label(uploaded_by.department)}. "
            f"El archivo contiene filas de: {foreign_labels}."
        )



def create_import_job(*, uploaded_file, uploaded_by=None) -> AttendanceImportJob:
    validate_excel_extension(uploaded_file.name)
    _validate_import_scope(uploaded_file=uploaded_file, uploaded_by=uploaded_by)
    return AttendanceImportJob.objects.create(
        uploaded_by=uploaded_by,
        source_file=uploaded_file,
        file_name=uploaded_file.name,
    )


def create_import_job_from_path(*, file_path: str, uploaded_by=None) -> AttendanceImportJob:
    with open(file_path, "rb") as handle:
        django_file = File(handle, name=file_path.split("/")[-1].split("\\")[-1])
        return create_import_job(uploaded_file=django_file, uploaded_by=uploaded_by)


def _serialize_value(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def _existing_raw_record_for_import(
    *,
    raw_full_name: str,
    raw_department: str,
    work_day,
    entry_at,
    exit_at,
) -> Optional[AttendanceRawRecord]:
    return (
        AttendanceRawRecord.objects.filter(
            normalized_full_name=normalize_text(raw_full_name),
            normalized_department=normalize_text(raw_department),
            work_day=work_day,
            entry_at=entry_at,
            exit_at=exit_at,
        )
        .order_by("created_at")
        .first()
    )


def _match_monitor(raw_record: AttendanceRawRecord) -> Tuple[Optional[Monitor], str]:
    mapped_department = _map_department(raw_record.normalized_department)
    if mapped_department is None:
        return None, "La dependencia del registro no coincide con ninguna dependencia configurada."
    matches = _match_monitor_queryset(
        raw_full_name=raw_record.raw_full_name,
        raw_department=raw_record.raw_department,
    )
    if matches.count() == 1:
        return matches.first(), ""
    if matches.count() > 1:
        return None, "Coincidencia ambigua por nombre y dependencia."
    return None, "No se encontró monitor por nombre y dependencia."


def _map_department(normalized_department: str) -> Optional[str]:
    mapping = {
        "monitores": "informatics_labs",
        "monitorias": "informatics_labs",
        "fisica": "physics",
        "monitores fisica": "physics",
        "physics": "physics",
        "Monitores Aulas de sistemas": "informatics_labs",
        "salas informatica": "informatics_labs",
        "informatica": "informatics_labs",
        "informatics labs": "informatics_labs",
        "monitores laboratorios": "electrical",
        "monitores laboratorio": "electrical",
        "laboratorios": "electrical",
        "electrica": "electrical",
        "electrical": "electrical",
    }
    return mapping.get(normalized_department)


@transaction.atomic
def reconcile_raw_record(raw_record: AttendanceRawRecord) -> AttendanceRawRecord:
    monitor, reason = _match_monitor(raw_record)
    if monitor:
        raw_record.monitor = monitor
        raw_record.reconciliation_status = ReconciliationStatusChoices.MATCHED
        raw_record.manual_review_reason = ""
    else:
        raw_record.reconciliation_status = ReconciliationStatusChoices.MANUAL_REVIEW
        raw_record.manual_review_reason = reason
        event_bus.publish(
            DomainEvent(
                name=ATTENDANCE_RECONCILIATION_FAILED,
                aggregate_id=str(raw_record.id),
                payload={
                    "raw_record_id": str(raw_record.id),
                    "reason": reason,
                    "raw_department": raw_record.raw_department,
                },
            )
        )
    raw_record.save(update_fields=["monitor", "reconciliation_status", "manual_review_reason", "updated_at"])
    return raw_record


@transaction.atomic
def assign_monitor_manually(*, raw_record: AttendanceRawRecord, monitor: Monitor, actor=None) -> AttendanceRawRecord:
    if actor is not None and not department_allowed(actor, monitor.department):
        raise ValidationError("No puedes conciliar monitores de otra dependencia.")
    raw_record.monitor = monitor
    raw_record.reconciliation_status = ReconciliationStatusChoices.MATCHED
    raw_record.manual_review_reason = ""
    raw_record.save(update_fields=["monitor", "reconciliation_status", "manual_review_reason", "updated_at"])
    return raw_record


def import_workbook(job: AttendanceImportJob) -> AttendanceImportJob:
    job.status = ImportJobStatusChoices.PROCESSING
    job.started_at = timezone.now()
    job.error_message = ""
    job.save(update_fields=["status", "started_at", "error_message", "updated_at"])

    try:
        workbook = load_workbook(job.source_file.path, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            raise ValueError("El archivo está vacío.")
        header_map = resolve_headers([str(item) for item in headers])

        imported_rows = 0
        failed_rows = 0
        total_rows = 0

        for row_number, row in enumerate(rows, start=2):
            if not any(value is not None and str(value).strip() for value in row):
                continue
            total_rows += 1
            payload = {str(headers[index]): _serialize_value(value) for index, value in enumerate(row)}
            # print(payload)
            try:
                worked_time = coerce_time(row[header_map["worked_time"]]) if "worked_time" in header_map else None
                entry_at = coerce_datetime(row[header_map["entry_at"]])
                work_day = f"{entry_at.year}-{entry_at.month}-{entry_at.day}"
                entry_at = f"{entry_at.hour}:{entry_at.minute}:{entry_at.second}"
                exit_at = coerce_datetime(
                    row[header_map["exit_at"]],
                    fallback_date=work_day or entry_at.date(),
                )
                exit_at = f"{exit_at.hour}:{exit_at.minute}:{exit_at.second}"
                work_day = coerce_date(work_day)
                entry_at = coerce_time(entry_at)
                exit_at = coerce_time(exit_at)
                # print(f"Tiempo: {worked_time}")
                # print(f"Entrada: {entry_at}")
                # print(f"Dia: {work_day}")
                # print(f"Salida: {exit_at}")
                raw_full_name = str(row[header_map["full_name"]]).strip()
                raw_department = str(row[header_map["department"]]).strip()
                existing_raw_record = _existing_raw_record_for_import(
                    raw_full_name=raw_full_name,
                    raw_department=raw_department,
                    work_day=work_day,
                    entry_at=entry_at,
                    exit_at=exit_at,
                )
                if existing_raw_record is not None:
                    logger.info(
                        "attendance_import_row_skipped_duplicate row=%s existing_raw_record=%s",
                        row_number,
                        existing_raw_record.id,
                    )
                    continue
                matched_monitors = _match_monitor_queryset(
                    raw_full_name=raw_full_name,
                    raw_department=raw_department,
                )
                monitor = matched_monitors.first() if matched_monitors.count() == 1 else None
                raw_record = AttendanceRawRecord.objects.create(
                    import_job=job,
                    row_number=row_number,
                    raw_full_name=raw_full_name,
                    raw_department=raw_department,
                    work_day = work_day,
                    entry_at = entry_at,
                    exit_at = exit_at,
                    worked_time = worked_time,
                    monitor = monitor,
                    # entry_at=timezone.make_aware(entry_at, timezone.get_current_timezone())
                    # if timezone.is_naive(entry_at)
                    # else entry_at,
                    # exit_at=timezone.make_aware(exit_at, timezone.get_current_timezone())
                    # if timezone.is_naive(exit_at)
                    # else exit_at,
                    raw_payload=payload,
                )
                reconcile_raw_record(raw_record)
                if raw_record.is_processable:
                    from apps.work_sessions.services import process_raw_record_to_session
                    process_raw_record_to_session(raw_record=raw_record)
                imported_rows += 1
            except Exception as exc:  # pragma: no cover - guarded by tests around service output
                failed_rows += 1
                logger.exception("attendance_import_row_failed row=%s error=%s", row_number, exc)

        job.total_rows = total_rows
        job.imported_rows = imported_rows
        job.failed_rows = failed_rows
        job.status = ImportJobStatusChoices.COMPLETED
        job.finished_at = timezone.now()
        job.save(
            update_fields=[
                "total_rows",
                "imported_rows",
                "failed_rows",
                "status",
                "finished_at",
                "updated_at",
            ]
        )
        event_bus.publish(
            DomainEvent(
                name=ATTENDANCE_IMPORTED,
                aggregate_id=str(job.id),
                payload={
                    "job_id": str(job.id),
                    "file_name": job.file_name,
                    "imported_rows": imported_rows,
                    "failed_rows": failed_rows,
                },
            )
        )
        return job
    except Exception as exc:
        job.status = ImportJobStatusChoices.FAILED
        job.finished_at = timezone.now()
        job.error_message = str(exc)
        job.save(update_fields=["status", "finished_at", "error_message", "updated_at"])
        logger.exception("attendance_import_failed job=%s error=%s", job.id, exc)
        raise
