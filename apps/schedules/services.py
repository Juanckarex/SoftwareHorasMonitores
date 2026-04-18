import re
from dataclasses import dataclass, field
from datetime import time
from typing import Iterable, Optional

from django.core.exceptions import ValidationError
from openpyxl import load_workbook

from apps.attendance.validators import validate_excel_extension
from apps.common.choices import UserRoleChoices
from apps.common.utils import normalize_text
from apps.monitors.models import Monitor
from apps.schedules.models import Schedule, ScheduleException
from apps.work_sessions.services import sync_sessions_for_exception_change


DAY_NAME_TO_WEEKDAY = {
    "lunes": Schedule.Weekday.MONDAY,
    "martes": Schedule.Weekday.TUESDAY,
    "miercoles": Schedule.Weekday.WEDNESDAY,
    "mircoles": Schedule.Weekday.WEDNESDAY,
    "jueves": Schedule.Weekday.THURSDAY,
    "viernes": Schedule.Weekday.FRIDAY,
    "sabado": Schedule.Weekday.SATURDAY,
    "sbado": Schedule.Weekday.SATURDAY,
    "domingo": Schedule.Weekday.SUNDAY,
}

DAY_HOUR_PATTERN = re.compile(
    r"(?P<day>[A-ZÁÉÍÓÚÜÑ]+)\s+(?P<start_hour>\d{1,2})(?::(?P<start_minute>\d{2}))?-(?P<end_hour>\d{1,2})(?::(?P<end_minute>\d{2}))?",
    re.IGNORECASE,
)


@dataclass
class ParsedScheduleBlock:
    weekday: int
    start_time: time
    end_time: time


@dataclass
class ScheduleImportResult:
    created: int = 0
    reactivated: int = 0
    processed_monitors: int = 0
    skipped_rows: int = 0
    missing_monitors: list[str] = field(default_factory=list)
    unauthorized_monitors: list[str] = field(default_factory=list)


def _row_label_value(row_values: list[str], target_label: str) -> Optional[str]:
    normalized_target = normalize_text(target_label).rstrip(":")
    for index, value in enumerate(row_values):
        normalized_value = normalize_text(value).rstrip(":")
        if normalized_value == normalized_target:
            for candidate in row_values[index + 1 :]:
                if candidate:
                    return candidate.strip()
    return None


def _build_time(hour_text: str, minute_text: Optional[str]) -> time:
    return time(hour=int(hour_text), minute=int(minute_text or 0))


def _parse_day_hour(value: str) -> Optional[ParsedScheduleBlock]:
    normalized_value = (value or "").strip().replace("?", "").replace("�", "")
    match = DAY_HOUR_PATTERN.search(normalized_value)
    if not match:
        return None
    normalized_day = normalize_text(match.group("day")).replace("?", "").replace("�", "")
    weekday = DAY_NAME_TO_WEEKDAY.get(normalized_day)
    if weekday is None:
        return None
    start_time = _build_time(match.group("start_hour"), match.group("start_minute"))
    end_time = _build_time(match.group("end_hour"), match.group("end_minute"))
    if end_time <= start_time:
        raise ValidationError("La hora final del bloque debe ser posterior a la hora inicial.")
    return ParsedScheduleBlock(weekday=weekday, start_time=start_time, end_time=end_time)


def _merge_schedule_blocks(blocks: Iterable[ParsedScheduleBlock]) -> list[ParsedScheduleBlock]:
    ordered_blocks = sorted(blocks, key=lambda block: (block.weekday, block.start_time, block.end_time))
    merged: list[ParsedScheduleBlock] = []
    for block in ordered_blocks:
        if not merged:
            merged.append(block)
            continue
        previous = merged[-1]
        if previous.weekday == block.weekday and block.start_time <= previous.end_time:
            merged[-1] = ParsedScheduleBlock(
                weekday=previous.weekday,
                start_time=previous.start_time,
                end_time=max(previous.end_time, block.end_time),
            )
            continue
        merged.append(block)
    return merged


def _save_schedule_block(*, monitor: Monitor, block: ParsedScheduleBlock, result: ScheduleImportResult) -> None:
    schedule, created = Schedule.objects.get_or_create(
        monitor=monitor,
        weekday=block.weekday,
        start_time=block.start_time,
        end_time=block.end_time,
        defaults={"is_active": True},
    )
    if created:
        result.created += 1
        return
    if not schedule.is_active:
        schedule.is_active = True
        schedule.save(update_fields=["is_active", "updated_at"])
        result.reactivated += 1


def _validate_schedule_import_scope(*, actor, monitor: Monitor) -> None:
    if actor is None or actor.role == UserRoleChoices.ADMIN:
        return
    if actor.role != UserRoleChoices.LEADER or not actor.department:
        raise ValidationError("Solo administradores y lideres pueden importar horarios.")
    if monitor.department != actor.department:
        raise ValidationError("Solo puedes importar horarios de tu propia dependencia.")


def _flush_monitor_blocks(
    *,
    monitor_code: Optional[str],
    monitor_name: Optional[str],
    blocks: list[ParsedScheduleBlock],
    result: ScheduleImportResult,
    actor=None,
) -> None:
    if not monitor_code or not blocks:
        return
    monitor = Monitor.objects.filter(codigo_estudiante=str(monitor_code).strip()).first()
    if monitor is None:
        result.missing_monitors.append(f"{monitor_name or 'Sin nombre'} ({monitor_code})")
        return
    try:
        _validate_schedule_import_scope(actor=actor, monitor=monitor)
    except ValidationError:
        result.unauthorized_monitors.append(
            f"{monitor.full_name} ({monitor.codigo_estudiante}) - {monitor.get_department_display()}"
        )
        return
    for block in _merge_schedule_blocks(blocks):
        _save_schedule_block(monitor=monitor, block=block, result=result)
    result.processed_monitors += 1


def upsert_schedule(*, monitor, weekday: int, start_time, end_time, is_active: bool = True) -> Schedule:
    schedule, _ = Schedule.objects.update_or_create(
        monitor=monitor,
        weekday=weekday,
        start_time=start_time,
        end_time=end_time,
        defaults={
            "is_active": is_active,
        },
    )
    return schedule


def _validate_exception_scope(*, actor, department) -> None:
    if actor.role == UserRoleChoices.ADMIN:
        return
    if not actor.department or department != actor.department:
        raise ValidationError("Solo puedes gestionar excepciones de tu propia dependencia.")


def save_schedule_exception(
    *,
    actor,
    instance: Optional[ScheduleException] = None,
    name: str,
    description: str,
    start_date,
    end_date,
    department,
    ignore_lateness: bool,
    approve_overtime: bool,
    is_active: bool,
):
    _validate_exception_scope(actor=actor, department=department)
    exception = instance or ScheduleException()
    previous_state = None
    if instance is not None and instance.pk:
        previous = ScheduleException.objects.get(pk=instance.pk)
        previous_state = {
            "start_date": previous.start_date,
            "end_date": previous.end_date,
            "department": previous.department,
        }

    exception.name = name
    exception.description = description
    exception.start_date = start_date
    exception.end_date = end_date
    exception.department = department
    exception.ignore_lateness = ignore_lateness
    exception.approve_overtime = approve_overtime
    exception.is_active = is_active
    exception.full_clean()
    exception.save()

    updated_sessions = sync_sessions_for_exception_change(
        current_exception=exception,
        previous_start_date=previous_state["start_date"] if previous_state else None,
        previous_end_date=previous_state["end_date"] if previous_state else None,
        previous_department=previous_state["department"] if previous_state else None,
    )
    return exception, updated_sessions


def delete_schedule_exception(*, actor, exception: ScheduleException) -> int:
    _validate_exception_scope(actor=actor, department=exception.department)
    previous_state = {
        "start_date": exception.start_date,
        "end_date": exception.end_date,
        "department": exception.department,
    }
    exception.delete()
    return sync_sessions_for_exception_change(
        previous_start_date=previous_state["start_date"],
        previous_end_date=previous_state["end_date"],
        previous_department=previous_state["department"],
    )


def import_schedules_from_workbook(*, uploaded_file, actor=None) -> ScheduleImportResult:
    validate_excel_extension(uploaded_file.name)
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    worksheet = workbook.active

    result = ScheduleImportResult()
    current_monitor_name: Optional[str] = None
    current_monitor_code: Optional[str] = None
    current_blocks: list[ParsedScheduleBlock] = []
    day_hour_column_index: Optional[int] = None

    for row in worksheet.iter_rows(values_only=True):
        row_values = [str(value).strip() if value is not None else "" for value in row]
        if not any(row_values):
            continue

        monitor_name = _row_label_value(row_values, "NOMBRE")
        if monitor_name is not None:
            _flush_monitor_blocks(
                monitor_code=current_monitor_code,
                monitor_name=current_monitor_name,
                blocks=current_blocks,
                result=result,
                actor=actor,
            )
            current_monitor_name = monitor_name
            current_monitor_code = None
            current_blocks = []
            day_hour_column_index = None
            continue

        monitor_code = _row_label_value(row_values, "CODIGO")
        if monitor_code is not None:
            current_monitor_code = monitor_code
            continue

        normalized_headers = [normalize_text(value).replace("/", "_") for value in row_values]
        if "dia_hora" in normalized_headers:
            day_hour_column_index = normalized_headers.index("dia_hora")
            continue

        if day_hour_column_index is None or day_hour_column_index >= len(row_values):
            continue

        day_hour_value = row_values[day_hour_column_index]
        if not day_hour_value:
            continue

        block = _parse_day_hour(day_hour_value)
        if block is None:
            result.skipped_rows += 1
            continue
        current_blocks.append(block)

    _flush_monitor_blocks(
        monitor_code=current_monitor_code,
        monitor_name=current_monitor_name,
        blocks=current_blocks,
        result=result,
        actor=actor,
    )
    return result
